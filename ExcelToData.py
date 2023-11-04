import openpyxl

def ExcelToData():
        # 打开Excel文件
        workbook = openpyxl.load_workbook('成绩表.xlsx')

        # 选择工作表
        worksheet = workbook['Sheet1']

        # 记录上一次读取的值
        previous_value = None

        # 保存相同值对应的科目和成绩的值
        data_dict = {}
        i = 1
        # 循环读取C列上的单元格内容
        for row in worksheet.iter_rows(min_row=3, min_col=3, max_col=3):
            
            for cell in row:
                
                # 获取C列的单元格值
                current_value = cell.value
                # 判断当前值与上一次值是否相同
                if current_value == previous_value:
                    # 获取相同值对应的D列和F列的值
                    d_column_value = worksheet.cell(row=cell.row, column=4).value
                    d_column_value = "".join(d_column_value.split())
                    f_column_value = worksheet.cell(row=cell.row, column=6).value
                    # 将D列和F列的值保存到字典中
                    data_dict.setdefault(current_value, {}).setdefault('成绩',{}).setdefault(d_column_value, f_column_value)
                else:
                    d_column_value = worksheet.cell(row=cell.row, column=4).value
                    d_column_value = "".join(d_column_value.split())
                    print(d_column_value)
                    f_column_value = worksheet.cell(row=cell.row, column=6).value
                    print(f_column_value)
                    # 将D列和F列的值保存到字典中
                    
                    data_dict[current_value]={'学号':worksheet.cell(row=cell.row, column=2).value}
                    data_dict.setdefault(current_value, {}).setdefault('成绩',{}).setdefault(d_column_value, f_column_value)
                    i = 1
                i = i+1
                #print(i)
                # 更新上一次值为当前值
                previous_value = current_value

        # 输出字典中的数据
            #print(data_dict)

        # 关闭Excel文件
        workbook.close()
        return data_dict

if __name__ == '__main__':
    ExcelToData()